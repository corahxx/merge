# handlers/station_merge_handler.py - 充电站多表合并

import re
import os
import tempfile
import importlib.util
import pandas as pd
from typing import List, Optional, Tuple
from io import BytesIO

def _get_operator_name_from_table_name(table_name: str) -> str:
    try:
        _dir = os.path.dirname(os.path.abspath(__file__))
        _path = os.path.join(_dir, "operator_name_rules.py")
        _spec = importlib.util.spec_from_file_location("operator_name_rules", _path)
        if _spec and _spec.loader:
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            return _mod.get_operator_name_from_table_name(table_name or "")
    except Exception:
        pass
    return ""

# 充电站标准列：以你提供的指标为准，再加 1.2 补全所需的 运营商名称、运营商类型；不包含充电桩相关列（如接口数量、接口1标准等）
STATION_STANDARD_COLUMNS = [
    "所属充电站编号", "充电站内部编号", "充电站名称", "充电站位置", "充电站投入使用时间",
    "充电站联系电话", "区县", "充电站所属区域分类", "充电站所属运营商", "充电站类型",
    "充电站集中度_转换", "服务时间", "站点内桩总数", "交流桩数量", "直流桩数量",
    "站点总装机功率", "交流桩总装机功率", "直流桩总装机功率", "备注",
    "运营商名称", "运营商类型",
]

COLUMN_SYNONYM_MAP = {
    "充电站编码": "所属充电站编号",
    "充电站内部编码": "充电站内部编号",
    "运营商": "运营商名称",
}


def _row_contains_text(row_series: pd.Series, *keywords: str) -> bool:
    text = " ".join(str(v) for v in row_series.astype(str).tolist())
    return any(kw in text for kw in keywords)


def detect_header_row(
    rows: List[pd.Series], table_name: str, max_rows: int = 3
) -> Tuple[Optional[int], Optional[str]]:
    """前 3 行内第一个包含「所属充电站编号」或「充电站编码」的行作为表头。"""
    header_keywords = ("所属充电站编号", "充电站编码")
    for i in range(min(len(rows), max_rows)):
        if _row_contains_text(rows[i], *header_keywords):
            return i, None
    return None, f"「{table_name}」由于无法确定表头暂未合并"


def clean_report_org_name(name: str) -> str:
    """与公共桩同一套规则：去 202512_公共桩_、_公共桩、附件一：及后缀。"""
    s = str(name)
    if "附件一：" in s:
        s = s.split("附件一：")[0]
    s = re.sub(r"^202512_公共桩_", "", s)
    s = s.replace("_公共桩", "")
    s = s.strip("_ .")
    if s.lower().endswith((".xlsx", ".xls", ".csv")):
        s = s[: s.rfind(".")]
    return s.strip() or name


def _get_first_n_rows(engine: str, sheet_name: str, file_bytes: bytes, n: int = 3) -> List[pd.Series]:
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=n, engine=engine)
    return [df.iloc[i] if i < len(df) else pd.Series() for i in range(n)]


def _sheet_has_content(
    file_bytes: bytes, sheet_name: str, engine: str, min_nonempty: int = 3
) -> Tuple[bool, Optional[str]]:
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=5, engine=engine)
        return (int(df.notna().sum().sum()) >= min_nonempty, None)
    except Exception as e:
        return (False, str(e))


def _find_target_sheet(file_bytes: bytes, engine: str) -> Tuple[Optional[str], bool, Optional[str]]:
    """
    充电站多 Sheet：1.1 优先 → 含「充电站」的 Sheet → 否则报错「多sheet表无法确定主表」。
    单 Sheet 直接返回。
    """
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes), engine=engine)
    except Exception as e:
        return None, False, f"打开工作簿失败: {e}"
    names = xl.sheet_names
    if not names:
        return None, False, "工作簿无工作表"
    if len(names) == 1:
        return names[0], False, None
    sheet_11 = next((n for n in names if "1.1" in n), None)
    if sheet_11:
        return sheet_11, True, None
    station_sheet = next((n for n in names if "充电站" in n), None)
    if station_sheet:
        return station_sheet, True, None
    return None, False, "多sheet表无法确定主表"


def _read_sheet_openpyxl_readonly(file_bytes: bytes, sheet_name: str, header_row: int) -> pd.DataFrame:
    from openpyxl import load_workbook
    from collections import defaultdict
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    col_row = rows[header_row]
    data_rows = rows[header_row + 1:]
    cnt = defaultdict(int)
    cols = []
    for c in col_row:
        key = str(c) if c is not None else ""
        cnt[key] += 1
        cols.append(key if cnt[key] == 1 else f"{key}.{cnt[key] - 1}")
    ncols = len(cols)
    data_rows = [list(r)[:ncols] + [None] * (ncols - len(r)) if len(r) != ncols else list(r) for r in data_rows]
    return pd.DataFrame(data_rows, columns=cols)


def _read_sheet_with_header(file_bytes: bytes, sheet_name: str, header_row: int, engine: str) -> pd.DataFrame:
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=header_row, engine=engine)
    except ValueError as e:
        if "wildcard" in str(e) or "filters" in str(e).lower():
            return _read_sheet_openpyxl_readonly(file_bytes, sheet_name, header_row)
        raise


def _find_header_row_in_sheet(
    file_bytes: bytes, sheet_name: str, engine: str, keyword: str, max_look: int = 5
) -> Tuple[Optional[int], Optional[str]]:
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=max_look, engine=engine)
    except Exception as e:
        return None, str(e)
    for i in range(len(df)):
        if _row_contains_text(df.iloc[i], keyword):
            return i, None
    return None, f"未找到含「{keyword}」的表头行"


def _enrich_from_12_13(
    df_11: pd.DataFrame,
    df_12: pd.DataFrame,
    df_13: pd.DataFrame,
    col_operator: str,
    col_manufacturer: str,
) -> pd.DataFrame:
    out = df_11.copy()
    op_id, op_name, op_type = "运营商编号", "运营商名称", "运营商类型"
    if op_id in df_12.columns and col_operator in out.columns:
        sub = df_12.drop_duplicates(op_id)
        if op_name in sub.columns:
            out["运营商名称"] = out[col_operator].map(sub.set_index(op_id)[op_name])
        if op_type in sub.columns:
            out["运营商类型"] = out[col_operator].map(sub.set_index(op_id)[op_type])
    man_id, man_name, man_type = "充电桩生产厂商编号", "充电桩生产厂商名称", "充电桩生产厂商类型"
    if man_id in df_13.columns and col_manufacturer in out.columns:
        sub = df_13.drop_duplicates(man_id)
        if man_name in sub.columns:
            out["充电桩生产厂商名称"] = out[col_manufacturer].map(sub.set_index(man_id)[man_name])
        if man_type in sub.columns:
            out["充电桩生产厂商类型"] = out[col_manufacturer].map(sub.set_index(man_id)[man_type])
    return out


def _align_to_standard(df: pd.DataFrame) -> pd.DataFrame:
    """对齐到充电站标准列；排除「(无列名)」列。"""
    rename = {}
    for c in df.columns:
        if str(c).strip() == "(无列名)":
            continue
        if c in STATION_STANDARD_COLUMNS:
            continue
        if c in COLUMN_SYNONYM_MAP:
            rename[c] = COLUMN_SYNONYM_MAP[c]
        elif str(c).strip() in COLUMN_SYNONYM_MAP:
            rename[c] = COLUMN_SYNONYM_MAP[str(c).strip()]
    df = df.rename(columns=rename)
    # 去掉 (无列名) 列
    df = df.loc[:, [c for c in df.columns if str(c).strip() != "(无列名)"]]
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated(keep="first")]
    for std in STATION_STANDARD_COLUMNS:
        if std not in df.columns:
            df[std] = pd.NA
    return df[STATION_STANDARD_COLUMNS]


def process_one_file(
    file_bytes: bytes,
    file_name: str,
    engine: str = "openpyxl",
) -> Tuple[Optional[pd.DataFrame], Optional[str], List[str]]:
    errors: List[str] = []
    ext = file_name[file_name.rfind("."):].lower() if "." in file_name else ""
    if ext == ".csv":
        try:
            df = pd.read_csv(BytesIO(file_bytes), header=0, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(BytesIO(file_bytes), header=0, encoding="gbk")
        df = _align_to_standard(df)
        report_org = clean_report_org_name(file_name)
        df.insert(0, "上报机构", report_org)
        return df, None, errors
    if ext not in (".xlsx", ".xls"):
        return None, f"「{file_name}」不支持的文件格式，已跳过", errors
    if ext == ".xls":
        engine = "xlrd"

    sheet_name, is_multi, sheet_error = _find_target_sheet(file_bytes, engine)
    if not sheet_name:
        msg = f"「{file_name}」未找到可用的工作表，已跳过"
        if sheet_error:
            msg += f"（{sheet_error}）"
        return None, msg, errors

    table_name = f"{file_name}（{sheet_name}）"
    rows = _get_first_n_rows(engine, sheet_name, file_bytes, n=3)
    header_row, err = detect_header_row(rows, table_name, max_rows=3)
    if err:
        return None, err, errors

    df = _read_sheet_with_header(file_bytes, sheet_name, header_row, engine)
    if df.empty or len(df.columns) == 0:
        return None, f"「{table_name}」表头解析后无数据，已跳过", errors

    # 1.2 运营商补全已取消；充电站标准列不含厂商列，不再读取 1.2/1.3

    df = _align_to_standard(df)
    report_org = clean_report_org_name(file_name)
    df.insert(0, "上报机构", report_org)
    op_name = _get_operator_name_from_table_name(file_name)
    if "运营商名称" in df.columns:
        df["运营商名称"] = op_name
    return df, None, errors


def merge_files(
    files: List[Tuple[str, bytes]],
    engine: str = "openpyxl",
) -> Tuple[Optional[pd.DataFrame], List[str], List[str], List[Tuple[str, int]]]:
    """接口与公共桩一致：(merged_df, success_list, error_list, row_counts)。"""
    merged: List[pd.DataFrame] = []
    success: List[str] = []
    errors: List[str] = []
    row_counts: List[Tuple[str, int]] = []
    for name, b in files:
        try:
            df, err, _ = process_one_file(b, name, engine)
        except Exception as e:
            errors.append(f"「{name}」合并时出错: {e}")
            continue
        if err:
            errors.append(err)
            continue
        if df is not None and len(df) > 0:
            merged.append(df)
            success.append(name)
            row_counts.append((name, len(df)))
    if not merged:
        return None, success, errors, []
    result = pd.concat(merged, ignore_index=True)
    return result, success, errors, row_counts


def merge_files_to_csv(
    files: List[Tuple[str, bytes]],
    engine: str = "openpyxl",
) -> Tuple[Optional[bytes], List[str], List[str], List[Tuple[str, int]]]:
    """大文件模式：合并为 CSV 字节流，返回 (csv_bytes, success_list, error_list, row_counts)。"""
    success: List[str] = []
    errors: List[str] = []
    row_counts: List[Tuple[str, int]] = []
    path = None
    try:
        first = True
        for name, b in files:
            try:
                df, err, _ = process_one_file(b, name, engine)
            except Exception as e:
                errors.append(f"「{name}」合并时出错: {e}")
                continue
            if err:
                errors.append(err)
                continue
            if df is None or len(df) == 0:
                continue
            if first:
                fd, path = tempfile.mkstemp(suffix=".csv")
                os.close(fd)
                df.to_csv(path, index=False, encoding="utf-8-sig", mode="w")
                first = False
            else:
                df.to_csv(path, index=False, encoding="utf-8-sig", mode="a", header=False)
            success.append(name)
            row_counts.append((name, len(df)))
        if path is None or not os.path.exists(path):
            return None, success, errors, []
        with open(path, "rb") as f:
            csv_bytes = f.read()
        return csv_bytes, success, errors, row_counts
    finally:
        if path and os.path.exists(path):
            try:
                os.unlink(path)
            except Exception:
                pass
