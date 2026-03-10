# handlers/table_merge_handler.py - 多运营商 Excel 表合并

import re
import tempfile
import os
import importlib.util
import pandas as pd
from typing import List, Dict, Optional, Tuple, Any
from io import BytesIO

# 运营商名称按表名提取（见 operator_name_rules.py）
def _get_operator_name_from_table_name(table_name: str) -> str:
    try:
        _dir = os.path.dirname(os.path.abspath(__file__))
        _path = os.path.join(_dir, "operator_name_rules.py")
        _spec = importlib.util.spec_from_file_location("operator_name_rules", _path)
        if _spec and _spec.loader:
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            return _mod.get_operator_name_from_table_name(table_name or "")
    except Exception:
        pass
    return ""

# 标准列名（在最少指标基础上增加 1.2/1.3 补全列：运营商类型、充电桩生产厂商名称、充电桩生产厂商类型）
STANDARD_COLUMNS = [
    "序号", "充电桩编号", "充电桩内部编号", "省份", "城市", "区县", "经度", "纬度", "经纬度标准",
    "充电桩类型", "充电桩所属区域分类", "所属充电站编号", "充电站内部编号", "充电站名称", "充电站位置",
    "充电站投入使用时间", "充电站所处道路属性", "充电站联系电话", "充电桩所属运营商", "电表号",
    "充电桩厂商编号", "充电桩型号", "充电桩属性", "充电桩生产日期", "服务时间", "桩型号是否获得联盟标识授权",
    "支付方式", "设备开通时间", "额定电压上限", "额定电压下限", "额定电流上限", "额定电流下限", "额定功率",
    "接口数量", "接口1标准", "接口2标准", "接口3标准", "接口4标准", "备注",
    "省份_中文", "城市_中文", "区县_中文", "充电桩类型_转换", "充电桩属性_转换", "充电桩所属运营商_转换",
    "充电桩厂商编号_转换", "入库时间", "运营商名称", "运营商类型", "充电桩内部编号_运营商名称",
    "充电桩生产厂商名称", "充电桩生产厂商类型",
]

# 同义列名映射：源列名 -> 标准列名（用于各运营商表头差异）
COLUMN_SYNONYM_MAP = {
    "充电桩编码": "充电桩编号",
    "充电桩内部编码": "充电桩内部编号",
    "充电站编码": "所属充电站编号",
    "充电站内部编码": "充电站内部编号",
    "运营商机构代码": "充电桩所属运营商",
    "接口标准": "接口1标准",
}


def _row_contains_text(row_series: pd.Series, *keywords: str) -> bool:
    """某行（单元格拼成字符串）是否包含任一关键词"""
    text = " ".join(str(v) for v in row_series.astype(str).tolist())
    return any(kw in text for kw in keywords)


def detect_header_row(
    rows: List[pd.Series], table_name: str, max_rows: int = 3
) -> Tuple[Optional[int], Optional[str]]:
    """
    在前 max_rows 行内查找表头行（0-based）。
    规则：1) 若第一行含「单位」「参考」「编码方法」任一个，则第二行含「充电桩编号」或「充电桩编码」则为表头；
    2) 否则第一行含「充电桩编号」或「充电桩编码」则为表头；3) 否则前3行内第一个含二者之一的为表头。
    表头阶段「编号」与「编码」等同。
    返回 (header_row, error_message)。若成功则 error_message 为 None。
    """
    # 表头关键词：编号与编码等同
    header_keywords = ("充电桩编号", "充电桩编码")
    row0_keywords = ("单位", "参考", "编码方法")

    if max_rows >= 2 and len(rows) >= 2:
        if _row_contains_text(rows[0], *row0_keywords):
            if _row_contains_text(rows[1], *header_keywords):
                return 1, None  # 第二行为表头，第一行不作为数据
    for i in range(min(len(rows), max_rows)):
        if _row_contains_text(rows[i], *header_keywords):
            return i, None
    return None, f"「{table_name}」由于无法确定表头暂未合并"


def clean_report_org_name(name: str) -> str:
    """
    清洗表名称用于填充「上报机构」：
    去掉 202512_公共桩_、_公共桩、附件一：及其后内容
    """
    s = str(name)
    if "附件一：" in s:
        s = s.split("附件一：")[0]
    s = re.sub(r"^202512_公共桩_", "", s)
    s = s.replace("_公共桩", "")
    s = s.strip("_ .")
    if s.lower().endswith((".xlsx", ".xls", ".csv")):
        s = s[: s.rfind(".")]
    return s.strip() or name


def _get_first_n_rows(engine: str, sheet_name: str, file_bytes: bytes, n: int = 3) -> List[pd.Series]:
    """读取指定 Sheet 的前 n 行（无表头，用于表头判定）"""
    df = pd.read_excel(
        BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=n, engine=engine
    )
    return [df.iloc[i] if i < len(df) else pd.Series() for i in range(n)]


def _sheet_has_content(
    file_bytes: bytes, sheet_name: str, engine: str, min_nonempty: int = 3
) -> Tuple[bool, Optional[str]]:
    """Sheet 是否有内容：至少 min_nonempty 个非空单元格（前几行）。返回 (是否通过, 异常信息若发生)。"""
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=5, engine=engine)
        return (int(df.notna().sum().sum()) >= min_nonempty, None)
    except Exception as e:
        return (False, str(e))


def _find_target_sheet(file_bytes: bytes, engine: str) -> Tuple[Optional[str], bool, Optional[str]]:
    """
    确定要读取的 Sheet：若存在名称含「1.1」的 Sheet 则返回该 Sheet（多 Sheet 模式）；
    否则返回第一个有内容的 Sheet（单 Sheet 模式）。
    返回 (sheet_name, is_multi_sheet, error_detail)。当 sheet_name 为 None 时 error_detail 可能含读表异常信息。
    """
    last_error: Optional[str] = None
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes), engine=engine)
    except Exception as e:
        return None, False, f"打开工作簿失败: {e}"
    names = xl.sheet_names
    if not names:
        return None, False, "工作簿无工作表"
    # 仅一个 Sheet 时直接使用，不判「有内容」
    if len(names) == 1:
        return names[0], False, None
    sheet_11 = next((n for n in names if "1.1" in n), None)
    first_non_empty = None
    for n in names:
        passed, err = _sheet_has_content(file_bytes, n, engine)
        if err:
            last_error = f"Sheet {n!r}: {err}"
        if passed:
            first_non_empty = n
            break
    if sheet_11:
        return sheet_11, True, None
    if first_non_empty is not None:
        return first_non_empty, False, None
    return None, False, last_error


def _read_sheet_openpyxl_readonly(file_bytes: bytes, sheet_name: str, header_row: int) -> pd.DataFrame:
    """用 openpyxl read_only 逐行读表，绕过 auto-filter 解析导致的 ValueError。"""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    # 表头行与数据行
    col_row = rows[header_row]
    data_rows = rows[header_row + 1 :]
    # 列名去重（与 pandas 行为一致：重复则 .1, .2, ...）
    from collections import defaultdict
    cnt = defaultdict(int)
    cols = []
    for c in col_row:
        key = str(c) if c is not None else ""
        cnt[key] += 1
        cols.append(key if cnt[key] == 1 else f"{key}.{cnt[key] - 1}")
    ncols = len(cols)
    # 统一每行长度为 ncols，避免 DataFrame 列数不一致报错
    data_rows = [list(r)[:ncols] + [None] * (ncols - len(r)) if len(r) != ncols else list(r) for r in data_rows]
    return pd.DataFrame(data_rows, columns=cols)


def _read_sheet_with_header(file_bytes: bytes, sheet_name: str, header_row: int, engine: str) -> pd.DataFrame:
    """读取 Sheet，指定行作为表头。openpyxl 解析筛选器报错时回退为 read_only 逐行读。"""
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=header_row, engine=engine)
    except ValueError as e:
        if "wildcard" in str(e) or "filters" in str(e).lower():
            return _read_sheet_openpyxl_readonly(file_bytes, sheet_name, header_row)
        raise


def _find_header_row_in_sheet(
    file_bytes: bytes, sheet_name: str, engine: str, keyword: str, max_look: int = 5
) -> Tuple[Optional[int], Optional[str]]:
    """在 Sheet 前 max_look 行内找第一个包含 keyword 的行作为表头行（0-based）。"""
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=max_look, engine=engine)
    except Exception as e:
        return None, str(e)
    for i in range(len(df)):
        if _row_contains_text(df.iloc[i], keyword):
            return i, None
    return None, f"未找到含「{keyword}」的表头行"


def _enrich_from_12_13(
    df_11: pd.DataFrame,
    df_12: pd.DataFrame,
    df_13: pd.DataFrame,
    col_operator: str,
    col_manufacturer: str,
) -> pd.DataFrame:
    """用 1.2、1.3 表补全 1.1 的运营商名称/类型、厂商名称/类型（map 方式，不改变列数）"""
    out = df_11.copy()
    op_id, op_name, op_type = "运营商编号", "运营商名称", "运营商类型"
    if op_id in df_12.columns and col_operator in out.columns:
        sub = df_12.drop_duplicates(op_id)
        if op_name in sub.columns:
            out["运营商名称"] = out[col_operator].map(sub.set_index(op_id)[op_name])
        if op_type in sub.columns:
            out["运营商类型"] = out[col_operator].map(sub.set_index(op_id)[op_type])
    man_id, man_name, man_type = "充电桩生产厂商编号", "充电桩生产厂商名称", "充电桩生产厂商类型"
    if man_id in df_13.columns and col_manufacturer in out.columns:
        sub = df_13.drop_duplicates(man_id)
        if man_name in sub.columns:
            out["充电桩生产厂商名称"] = out[col_manufacturer].map(sub.set_index(man_id)[man_name])
        if man_type in sub.columns:
            out["充电桩生产厂商类型"] = out[col_manufacturer].map(sub.set_index(man_id)[man_type])
    return out


def _align_to_standard(df: pd.DataFrame) -> pd.DataFrame:
    """将 DataFrame 列对齐到标准列名（同义映射 + 缺列填 NaN），不包含上报机构"""
    rename = {}
    for c in df.columns:
        if c in STANDARD_COLUMNS:
            continue
        if c in COLUMN_SYNONYM_MAP:
            rename[c] = COLUMN_SYNONYM_MAP[c]
        elif str(c).strip() in COLUMN_SYNONYM_MAP:
            rename[c] = COLUMN_SYNONYM_MAP[str(c).strip()]
    df = df.rename(columns=rename)
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated(keep="first")]
    for std in STANDARD_COLUMNS:
        if std not in df.columns:
            df[std] = pd.NA
    return df[STANDARD_COLUMNS]


def process_one_file(
    file_bytes: bytes,
    file_name: str,
    engine: str = "openpyxl",
) -> Tuple[Optional[pd.DataFrame], Optional[str], List[str]]:
    """
    处理单个文件，返回 (合并用的 DataFrame, 错误信息, 警告/跳过列表)。
    若成功则 error 为 None；若跳过则 df 为 None 且 error 为跳过原因。
    """
    errors: List[str] = []
    ext = file_name[file_name.rfind(".") :].lower() if "." in file_name else ""
    if ext == ".csv":
        try:
            df = pd.read_csv(BytesIO(file_bytes), header=0, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(BytesIO(file_bytes), header=0, encoding="gbk")
        df = _align_to_standard(df)
        report_org = clean_report_org_name(file_name)
        df.insert(0, "上报机构", report_org)
        return df, None, errors
    if ext not in (".xlsx", ".xls"):
        return None, f"「{file_name}」不支持的文件格式，已跳过", errors
    if ext == ".xls":
        engine = "xlrd"

    sheet_name, is_multi, sheet_error = _find_target_sheet(file_bytes, engine)
    if not sheet_name:
        msg = f"「{file_name}」未找到可用的工作表，已跳过"
        if sheet_error:
            msg += f"（{sheet_error}）"
        return None, msg, errors

    table_name = f"{file_name}（{sheet_name}）"
    rows = _get_first_n_rows(engine, sheet_name, file_bytes, n=3)
    header_row, err = detect_header_row(rows, table_name, max_rows=3)
    if err:
        return None, err, errors

    df = _read_sheet_with_header(file_bytes, sheet_name, header_row, engine)
    if df.empty or len(df.columns) == 0:
        return None, f"「{table_name}」表头解析后无数据，已跳过", errors

    # 1.2 运营商补全已取消；仅保留 1.3 厂商补全
    if is_multi:
        xl = pd.ExcelFile(BytesIO(file_bytes), engine=engine)
        df_13 = None
        for sn in xl.sheet_names:
            if "1.3" in sn:
                h13, err13 = _find_header_row_in_sheet(file_bytes, sn, engine, "充电桩生产厂商名称")
                if err13 is None and h13 is not None:
                    try:
                        df_13 = _read_sheet_with_header(file_bytes, sn, h13, engine)
                    except Exception:
                        pass
        col_man = "充电桩厂商编号"
        if df_13 is not None:
            df = _enrich_from_12_13(
                df,
                pd.DataFrame(),
                df_13,
                "充电桩所属运营商",
                col_man,
            )

    df = _align_to_standard(df)
    report_org = clean_report_org_name(file_name)
    df.insert(0, "上报机构", report_org)
    op_name = _get_operator_name_from_table_name(file_name)
    if "运营商名称" in df.columns:
        df["运营商名称"] = op_name
    return df, None, errors


def merge_files(
    files: List[Tuple[str, bytes]],
    engine: str = "openpyxl",
) -> Tuple[Optional[pd.DataFrame], List[str], List[str], List[Tuple[str, int]]]:
    """
    合并多个文件。files = [(file_name, file_bytes), ...]
    返回 (merged_df, success_list, error_list, row_counts)。
    row_counts 为 [(文件名, 行数), ...]，顺序与 success_list 一致；多 Sheet 时行数仅统计 1.1 表。
    """
    merged: List[pd.DataFrame] = []
    success: List[str] = []
    errors: List[str] = []
    row_counts: List[Tuple[str, int]] = []
    for name, b in files:
        try:
            df, err, _ = process_one_file(b, name, engine)
        except Exception as e:
            errors.append(f"「{name}」合并时出错: {e}")
            continue
        if err:
            errors.append(err)
            continue
        if df is not None and len(df) > 0:
            merged.append(df)
            success.append(name)
            row_counts.append((name, len(df)))
    if not merged:
        return None, success, errors, []
    result = pd.concat(merged, ignore_index=True)
    return result, success, errors, row_counts


def merge_files_to_csv(
    files: List[Tuple[str, bytes]],
    engine: str = "openpyxl",
) -> Tuple[Optional[bytes], List[str], List[str], List[Tuple[str, int]]]:
    """
    大文件模式：合并多个文件并直接输出为 CSV 字节流，不返回 DataFrame。
    返回 (csv_bytes, success_list, error_list, row_counts)。无合并数据时 csv_bytes 为 None。
    """
    success: List[str] = []
    errors: List[str] = []
    row_counts: List[Tuple[str, int]] = []
    path = None
    try:
        first = True
        for name, b in files:
            try:
                df, err, _ = process_one_file(b, name, engine)
            except Exception as e:
                errors.append(f"「{name}」合并时出错: {e}")
                continue
            if err:
                errors.append(err)
                continue
            if df is None or len(df) == 0:
                continue
            if first:
                fd, path = tempfile.mkstemp(suffix=".csv")
                os.close(fd)
                df.to_csv(path, index=False, encoding="utf-8-sig", mode="w")
                first = False
            else:
                df.to_csv(path, index=False, encoding="utf-8-sig", mode="a", header=False)
            success.append(name)
            row_counts.append((name, len(df)))
        if path is None or not os.path.exists(path):
            return None, success, errors, []
        with open(path, "rb") as f:
            csv_bytes = f.read()
        return csv_bytes, success, errors, row_counts
    finally:
        if path and os.path.exists(path):
            try:
                os.unlink(path)
            except Exception:
                pass
